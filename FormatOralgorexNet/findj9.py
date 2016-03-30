import re,os
import openpyxl
from openpyxl.styles import Font, Fill,Color
from openpyxl.cell import get_column_letter
from openpyxl import Workbook
from openpyxl.styles.colors import RED , BLUE , GREEN  

path1 =  r'F:\XYZ\PROJECTS\SAC\SACSIMULATOR\SACSIMV4.NET'
path2 = r'C:\Users\michaelbru\Documents\Visual Studio 2013\Projects\FormatOralgorexNet\FormatOralgorexNet\AdaptorMAP.xlsx'


def openXl(bookname):

            #try: 
            #    os.remove(bookname) 
                
            #except: 
            #    pass
# Create a workbook and add a worksheet.
        if os.path.exists( bookname ):
           return openpyxl.load_workbook(bookname,guess_types=False)
        else:
            raise ValueError

def saveXl(wb,bookname):
    wb.save(bookname)



#open netlist
data = open(path1,'r').read()
dataList = data.split('\n')
#dataList = data.split('\n')


#format netlist as dictionary
newData = {}
for i in range(len(dataList)):
    if re.search('^\S',dataList[i]):
        j = i+1
        newData[dataList[i]] =[]
        while  j < len(dataList): 
            if re.search('^\s',dataList[j]):
                res = dataList[j].lstrip()
                if res!='':
                    if res[-1]==',':
                     res = res[:res.find('(')-1]+res[res.find(')')+1:-1]
                    else:
                        res = res[:res.find('(')-1]+res[res.find(')')+1:]
                    newData[dataList[i]].append(res)
                j+=1
            else:
                break
#print(newData)






# find all nets with J9
#########################################################

wb = Workbook()
ws = wb.active
ws.title = 'J9'
#create new dictionary j9 from  newData 
j9 = {}
for k in newData:
    for v in newData[k]:
        if v.find('J9')!=-1 and v[0]!='U':
            j9[k] = v
#j9 = [ k  for x in newData ]
print(j9)

# assign j9 to worksheet
i=1
for k in j9 :
#    for v in j9[k]:
        ws['A%s'%i] = k
        ws['B%s'%i] = j9[k]
        i+=1

#################################################
#open map xml in MALE worksheet
wbMap =  openXl(path2)
wsMap = wbMap['MALE']
# assign location of line to start from it 
startLine = 5
# assign location of columns
fromSacColumn  = 'A'
pcbColumn = 'C'
 
#create new list with PCB pinout - column C   
mapArr = []
for i in range(50):
    mapArr.append((i+1, wsMap['C%s'%(i+startLine)].value))
#for i in range(1,51):
#    print(i,mapArr[i-1])
######################################################

# do mapping


res = {}
#for i in range(1,51):
count =1
for k in j9:
    toChange = int(j9[k][j9[k].find('-')+1:])   # the last pin number(J9-10)
    for i,pcb in mapArr:
        if toChange == pcb:
            res[k] =  j9[k].replace(str(toChange),str(i))
            ws["C%s"%count] = k
            ws["D%s"%count] = res[k]
            count+=1
    #res[k] =  j9[k].replace(str(toChange),str(mapArr[ toChange]))
    #if i == int(j9[k][j9[k].find('-')+1:]):
    #    j9[k][j9[k].find('-')+1:]
    #    print((j9[k][j9[k].find('-')+1:]))
    #print(toChange,mapArr[ toChange])
print(res)
saveXl(wb,r'C:\Users\michaelbru\Documents\Visual Studio 2013\Projects\FormatOralgorexNet\FormatOralgorexNet\j9.xlsx')
