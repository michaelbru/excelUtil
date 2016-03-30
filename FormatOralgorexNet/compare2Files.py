import re,os
import openpyxl
from openpyxl.styles import Font, Fill,Color
from openpyxl.cell import get_column_letter
from openpyxl import Workbook
from openpyxl.styles.colors import RED , BLUE , GREEN  


def openXl(bookname):

            #try: 
            #    os.remove(bookname) 
                
            #except: 
            #    pass
# Create a workbook and add a worksheet.
        if os.path.exists( bookname ):
           return openpyxl.load_workbook(bookname,guess_types=False)
        else:
            raise ValueError

def saveXl(wb,bookname):
    wb.save(bookname)


def proccesXl(bookname1,bookname2):
    try:
        wb1 =openXl(bookname1)
        wb2 =openXl(bookname2)
    except ValueError as ve:
        print( "Probably the file name is wrong " )

    wb1Sheets = ['J7_CABLE']#['MOTOR CURRENT','EHV_CURR_IN&EHV_COIL','BYPASS','SOV & DISC','Brakes','EHV_LVDT & POS_LVDT']
    wb2Sheets = ['MOTOR CURRENT','EHV_CURR_IN&EHV_COIL','BYPASS','SOV & DISC','Brakes','EHV_LVDT & POS_LVDT','INPUT_TO_SIM_ANALOG_DC','SIM OUTPUT ANALOG DC VOLT','FrontPanel']
    wb = Workbook()
    ws = wb.active
    ws.title = "Nets"
   # f = open(r'C:\Users\michaelbru\Documents\Visual Studio 2013\Projects\FormatOralgorexNet\FormatOralgorexNet\rec.txt','a')
    c = 1
    ##open sheet 
    # search for every sheet in two workbooks and find lines with the same nets
    # create new workbook and write them in it
    for wb1Sheet in wb1Sheets:
        print(wb1Sheet+'~~~~~~~~')
      
        for  wb2Sheet in wb2Sheets:
            l=100  #compare 100 lines
            print(wb2Sheet)
            for i in range(1,100):
                for j in range(1,100):
                    #print(j)
                    r =  wb1[wb1Sheet][chr(ord('B'))+str(i)].value
                    tsted = r[1:-1] if type(r)==str else ''
                   
                    src = wb2[wb2Sheet][chr(ord('A'))+str(j)].value
                    
                    if  wb2[wb2Sheet][chr(ord('A'))+str(j)].value!=None:
                       if wb2[wb2Sheet][chr(ord('A'))+str(j)].value!='NET' :
                        if type(src) == str and src == tsted or src==r :
                           
                            ws['A%s'%c].value =  wb1[wb1Sheet][chr(ord('B'))+str(i)].value
                            ws['B%s'%c].value =  wb1[wb1Sheet][chr(ord('C'))+str(i)].value
                            ws['C%s'%c].value =  wb1[wb1Sheet][chr(ord('D'))+str(i)].value

                            ws['D%s'%c].value = wb1Sheet
                            print(wb1[wb1Sheet][chr(ord('D'))+str(j)].value)
                          

                            ws['F%s'%c].value =  wb2[wb2Sheet][chr(ord('A'))+str(j)].value
                            ws['G%s'%c].value =  wb2[wb2Sheet][chr(ord('B'))+str(j)].value
                            ws['H%s'%c].value =  wb2[wb2Sheet][chr(ord('C'))+str(j)].value
                            ws['I%s'%c].value =  wb2[wb2Sheet][chr(ord('D'))+str(j)].value
                            ws['J%s'%c].value =  wb1Sheet
                            print(wb2[wb2Sheet][chr(ord('A'))+str(j)].value)
                            c+=1
                            print(c)

   # f.close()
    saveXl(wb,r'F:\xyz\projects\Sac\resultsComparedNets.xlsx')



def main():
    proccesXl(  r'F:\xyz\projects\Sac\SacSimulator\SacSimulator\MLT_CABLE_IliaOnly.xlsx',r'F:\xyz\projects\Sac\wiringBK2.xlsx'
            )
    #replace(r'C:\Users\michaelbru\Documents\Visual Studio 2013\Projects\FormatOralgorexNet\FormatOralgorexNet\wiring.xlsx',
    #        r'C:\Users\michaelbru\Documents\Visual Studio 2013\Projects\FormatOralgorexNet\FormatOralgorexNet\resultsNets.xlsx')
    # sort(r'C:\Users\michaelbru\Documents\Visual Studio 2013\Projects\FormatOralgorexNet\FormatOralgorexNet\resultsNets.xlsx')
if __name__ == "__main__":
    main()