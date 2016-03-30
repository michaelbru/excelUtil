import re
import openpyxl
from openpyxl.styles import Font, Fill,Color
from openpyxl.cell import get_column_letter

from openpyxl.styles.colors import RED , BLUE , GREEN  
import getopt, sys,os



def openXl(bookname):

            #try: 
            #    os.remove(bookname) 
                
            #except: 
            #    pass
# Create a workbook and add a worksheet.
        if os.path.exists( bookname ):
           return openpyxl.load_workbook(bookname,guess_types=False)
        else:
            raise ValueError

def saveXl(wb,bookname):
    wb.save(bookname)

def proccesXl(bookname,col,pattern,word):
    try:
        wb =openXl(bookname)
    except ValueError as ve:
        print( "Probably the file name is wrong " )
    ws = wb.active
    c = col.upper()
    count = 1
    while ws[c+str(count)].value:
        if ws[c+str(count)].value.startswith(pattern) :
            ws[chr(ord(c)+1)+str(count)] = word
        count+=1

    saveXl(wb,bookname)

def main():
    try:
        opts, args = getopt.getopt(sys.argv[1:], "h:f:c:w:p:" , ["help", "file=","col=","word=","pattr="])
    except getopt.GetoptError as err:
        # print help information and exit:
        print(err) # will print something like "option -a not recognized"
        usage()
        sys.exit(2)

    file = ''
    col = 'A'
    pattern = ''
    word = ''

    for o, a in opts:
        if o in ("-c","--col"):
            col = a
            print('col=%s'% a)
        elif o in ("-h", "--help"):
            usage()
            sys.exit()
        elif o in ("-f", "--file"):
            file = a
            print('file=%s'% a)
        elif o in ("-p", "--pattr"):
            pattern = a
            print('pattr=%s'% a)
        elif o in ("-w", "--word"):
            word = a
            print('word=%s'% a)
        else:
            assert False, "unhandled option"
    # ...
    try:
        proccesXl(file,col,pattern,word)
    except Exception as ex:
        print(ex)
        

if __name__ == "__main__":
    main()