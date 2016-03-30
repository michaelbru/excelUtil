import re,os
import openpyxl
from openpyxl.styles import Font, Fill,Color
from openpyxl.cell import get_column_letter
from openpyxl import Workbook
from openpyxl.styles.colors import RED , BLUE , GREEN  

path1 =  r'F:\XYZ\PROJECTS\SAC\SACSIMULATOR\SACSIMV4.NET'
#path2 = r'C:\Users\michaelbru\Documents\Visual Studio 2013\Projects\FormatOralgorexNet\FormatOralgorexNet\AdaptorMAP.xlsx'


def openXl(bookname):

            #try: 
            #    os.remove(bookname) 
                
            #except: 
            #    pass
# Create a workbook and add a worksheet.
        if os.path.exists( bookname ):
           return openpyxl.load_workbook(bookname,guess_types=False)
        else:
            raise ValueError

def saveXl(wb,bookname):
    wb.save(bookname)


#open netlist
data = open(path1,'r').read()
dataList = data.split('\n')
#dataList = data.split('\n')
wb = Workbook()
ws = wb.active
ws.title = 'J6'


#format netlist as dictionary
newData = {}
for i in range(len(dataList)):
    if re.search('^\S',dataList[i]): # find line that doesnt begin from blank
        j = i+1
        newData[dataList[i]] =[] # open dictionary for this net
        while  j < len(dataList): 
            if re.search('^\s',dataList[j]): # find 
                res = dataList[j].lstrip()
                if res!='':
                    if res[-1]==',':
                     res = res[:res.find('(')-1]+res[res.find(')')+1:-1] #filter line from string inside brackets including brackets, and commas
                    else:
                        res = res[:res.find('(')-1]+res[res.find(')')+1:] #filter line from string inside brackets including brackets
                    newData[dataList[i]].append(res)
                j+=1
            else:
                break
# find J6 
j6 = {}
for k in newData:
    for v in newData[k]:
        if v.find('J6')!=-1 and v[0]!='U':
            j6[k] = v
#j9 = [ k  for x in newData ]
print(j6)


mapArr = [1,3,5,7,9,11,13,15,17,19,21,23,25,27,29,31,33,35,37,2,4,6,8,10,12,14,16,18,20,22,24,26,28,30,32,34,36]

res={}
for k in j6:
    v=j6[k]
    i = int(v[v.find('-')+1:])
    #print(i)
    res[k]='P6-%s'% (mapArr.index(i)+1)


ws['A1'].value = 'NET'
ws['B1'].value = 'PCB P6'
ws['C1'].value =  'Connector J6'
i=2
for k in j6:

    print(k,j6[k], '=>',res[k])
    ws['A%s'%i].value = k
    ws['B%s'%i].value = j6[k]
    ws['C%s'%i].value = res[k]
    i+=1
saveXl(wb,r'F:\xyz\projects\Sac\j6.xlsx')
