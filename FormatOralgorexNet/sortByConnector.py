import re,os
import openpyxl
from openpyxl.styles import Font, Fill,Color
from openpyxl.cell import get_column_letter
from openpyxl import Workbook
from openpyxl.styles.colors import RED , BLUE , GREEN  


path = r'F:\xyz\projects\Sac\wiringBK2.xlsx'


def openXl(bookname):

            #try: 
            #    os.remove(bookname) 
                
            #except: 
            #    pass
# Create a workbook and add a worksheet.
        if os.path.exists( bookname ):
           return openpyxl.load_workbook(bookname,guess_types=False)
        else:
            raise ValueError

def saveXl(wb,bookname):
    wb.save(bookname)



#sheets  = ['BYPASS']
sheets  = ['MOTOR CURRENT','EHV_CURR_IN&EHV_COIL','BYPASS','SOV & DISC','Brakes','EHV_LVDT & POS_LVDT','INPUT_TO_SIM_ANALOG_DC','SIM OUTPUT ANALOG DC VOLT','FrontPanel']
#sheets  = ['MOTOR_CURRENT','EHV_CURR_IN&EHV_COIL','BYPASS','SOV&DISC','Brakes','EHV_LVDT&POS_LVDT','INPUT_TO_SIM_ANALOG_DC','SIM_OUTPUT_ANALOG_DC_VOLT','FrontPanel']
connectors = ['P1','P2','P5','P6','P7','P8','P9']
created = []
wb = openXl(path)

wb1 = Workbook()

for c in connectors:
    if c not in created:
                created.append(c)
                ws2 = wb1.create_sheet(title=c)
    i = 1            
    for sheet in sheets:
        #print(sheet)
        ws = wb[sheet]
        
       
        for cell in range(1,100):
           
            if str(ws['B%s'%cell].value).startswith(c):

                ws2['A%s'%i].value = ws['A%s'%cell].value
                ws2['B%s'%i].value = ws['B%s'%cell].value
                ws2['C%s'%i].value = ws['C%s'%cell].value
                ws2['D%s'%i].value = ws['D%s'%cell].value
                ws2['E%s'%i].value = 'From %s'%sheet
             #   ws2['A%s'%i].value = ws['B%s'%cell].value
                if  sheet == 'INPUT TO SIM ANALOG DC':
                    print(i,c,ws2['A%s'%i].value,ws2['B%s'%i].value)

                i+=1
              



saveXl(wb1,r'F:\xyz\projects\Sac\wiring10.xlsx')